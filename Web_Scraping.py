#!/usr/bin/python3.6

# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import pandas as pd
import sys, getopt
import os
import time

# Work Space

Work_Space_0 = "D:\\self-knowledge\\Trieu"
Work_Space_1 = "D:/self-knowledge/Trieu"
# Here is the source of html files for ConCung
source_dir_ConCung_0 = Work_Space_0 + "\\example_ConCung"
source_dir_ConCung_1 = Work_Space_1 + "/example_ConCung"

# Here is the source of html files for BiBo
source_dir_BiBo_0 = Work_Space_0 + "\\example_BiBo"
source_dir_BiBo_1 = Work_Space_1 + "/example_BiBo"

# Here is the source of html files for KidsPlaza
source_dir_KidsPlaza_0 = Work_Space_0 + "\\example_KidsPlaza"
source_dir_KidsPlaza_1 = Work_Space_1 + "/example_KidsPlaza"

# Here is the source of html files for ShopTreTho
source_dir_ShopTreTho_0 = Work_Space_0 + "\\example_ShopTreTho"
source_dir_ShopTreTho_1 = Work_Space_1 + "/example_ShopTreTho"

# Here is the source of html files for ShopTreTho
source_dir_TutiCare_0 = Work_Space_0 + "\\example_TutiCare"
source_dir_TutiCare_1 = Work_Space_1 + "/example_TutiCare"

# Here is the source of html files for BXH
source_dir_BXH_0 = Work_Space_0 + "\\example_BXH"
source_dir_BXH_1 = Work_Space_1 + "/example_BXH"

# Here is the source of html files for VinMart
source_dir_VinMart_0 = Work_Space_0 + "\\example_VinMart"
source_dir_VinMart_1 = Work_Space_1 + "/example_VinMart"

# Here is the source of html files for KiotViet
source_dir_KiotViet_0 = Work_Space_0 + "\\example_KiotViet"
source_dir_KiotViet_1 = Work_Space_1 + "/example_KiotViet"

# =============================================================================
# This function is used to get list of files from a directory
#
# Parameters in:
#	@dir 	: the directory of source file (include html files)
#
# Return: a list of file names
# =============================================================================
def get_list_of_file(dir):
	return os.listdir(dir)

# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://concung.com/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#
# Return:	NULL
# =============================================================================
def ConCung_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 2 columns: Product Name and Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_ConCung_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("div", {"class": "product-item"}):
			try:
				# Get product name
				product_name = item.a['title']

				# Get price of product
				Price = item.find("strong",{"class":"product-price"}).contents[0]

				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name
				ws.cell(row=pointer, column=2).value = Price
			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass
			pointer = pointer + 1

	wb.save(out_file)


# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://bibomart.com.vn/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
def BiBo_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 2 columns: Product Name and Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	ws['C1'] = "QR Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_BiBo_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("div", {"class": "product details product-item-details"}):
			try:
				# Get product name
				product_name = item.find("a",{"class":"product-item-link"}).contents[0]

				# Get price of product
				Price1 = item.find("span",{"class":"price"}).contents[0]
				try:
					Price2 = item.find("div",{"class":"vnpay-price-items"}).find("span",{"class":"price"}).contents[0].contents[0]
				except:
					Price2 ="-"
				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name.strip('\n')
				ws.cell(row=pointer, column=2).value = Price1
				ws.cell(row=pointer, column=3).value = Price2
			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass

			pointer = pointer + 1

	wb.save(out_file)


# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://www.kidsplaza.vn/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
def KidsPlaza_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 2 columns: Product Name and Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	ws['C1'] = "QR Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_KidsPlaza_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("li", {"itemprop": "itemListElement"}):
			try:
				# Get product name
				product_name = item.find("span",{"itemprop":"name"}).contents[0]
				# Get price of product
				Price1 = item.find("span",{"itemprop":"price"}).contents[0]
				try:
					Price2 = item.find("div",{"class":"vnpay-price-items"}).find("span",{"class":"price"}).contents[0]
				except:
					Price2 ="-"
				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name.strip('\n')
				ws.cell(row=pointer, column=2).value = Price1
				ws.cell(row=pointer, column=3).value = Price2
			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass

			pointer = pointer + 1

	wb.save(out_file)



# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://shoptretho.com.vn/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
def ShopTreTho_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 2 columns: Product Name and Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	ws['C1'] = "Old Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_ShopTreTho_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("div", {"class": "product"}):
			try:
				product_name = item.find("h3",{"class":"name_pro"}).a.contents[0]
				Price1 = item.find("span",{"class":"price_item"}).contents[0]
				try:
					Price2 = item.find("span",{"class":"old_price"}).contents[0]
				except:
					Price2 ="-"

				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name.strip('\n')
				ws.cell(row=pointer, column=2).value = Price1
				ws.cell(row=pointer, column=3).value = Price2

			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass
			pointer = pointer + 1

	wb.save(out_file)


# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://www.tuticare.com/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
def TutiCare_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 3 columns: Product Name, Price and Old Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	ws['C1'] = "Old Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_TutiCare_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("div", {"class": "p_container"}):
			try:
				product_name = item.find("a",{"class":"p-name"}).contents[0]
				Price1 = item.find("span",{"class":"p-price-fomat"}).contents[0]
				try:
					Price2 = item.find("span",{"class":"p-old-price"}).contents[0]
				except:
					Price2 ="-"

				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name.strip('\n')
				ws.cell(row=pointer, column=2).value = Price1
				ws.cell(row=pointer, column=3).value = Price2

			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass
			pointer = pointer + 1

	wb.save(out_file)



# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://www.bachhoaxanh.com/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
def BXH_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 3 columns: Product Name, Price and Old Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	ws['C1'] = "Old Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_BXH_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("li"):
			try:
				product_name = item.find("div",{"class":"product-name"}).contents[0]
				Price1 = item.find("div",{"class":"price"}).strong.contents[0]
				try:
					Price2 = item.find("div",{"class":"price"}).span.contents[0]
				except:
					Price2 ="-"

				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name.strip('\n')
				ws.cell(row=pointer, column=2).value = Price1
				ws.cell(row=pointer, column=3).value = Price2

			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass
			pointer = pointer + 1

	wb.save(out_file)

# cat banh-quy-walens.1 | grep "product-name\|strong" | sed '/nav>/d' | sed '/important/d' | sed '/<a/d' | sed '/<em/d' | sed 's/<div class="product-name">//g' | sed 's/<\/div>//g' | sed 's/<strong>/MYKEYWORD/g' | sed 's/<\/strong>//g' | sed 's/  //g' | sed -z 's/\r\nMYKEYWORD/\t/g' | sed '/MYKEYWORD/d' >> All_Price.txt 


# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://vinmart.com/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
# def VinMart_get_and_write_data(list_file, out_file):

# 	T.B.D

# =============================================================================
# This function is used to get data (product name and price) from html files of
# https://www.kiotviet.vn/ to write to excel file.
#
# Parameters in:
#	@list_file: list of file name with html format.
#	@out_file:	name of excel file that you want to export.
#	
# Return:	NULL
# =============================================================================
def KiotViet_get_and_write_data(list_file, out_file):

	wb = Workbook()
	ws = wb.active
	# Write 3 columns: Product Name, Price and Old Price to excel file
	ws['A1'] = "Product Name"
	ws['B1'] = "Price"
	ws['C1'] = "Old Price"
	pointer = 2

	# Loop all file in list of files
	for file in list_file:

		soup = BeautifulSoup(open(source_dir_KiotViet_0+"\\"+file,'r',encoding='utf8',errors="ignore"), 'lxml')
		# Loop all product-tags in a file
		for item in soup.find_all("div",{"class":"kv-product-list-item--content"}):
			try:
				product_name = item.find("h6",{"class":"kv-title"}).contents[0]
				Price1 = item.find("p",{"class":"regular"}).contents[0]

				# Write the data to excel file
				ws.cell(row=pointer, column=1).value = product_name.strip('\n')
				ws.cell(row=pointer, column=2).value = Price1

			except:
				pointer = pointer - 1
				print("ERROR occurs - keep doing")
				pass
			pointer = pointer + 1

	wb.save(out_file)


# =============================================================================
# This function is used to remove dupplicated rows in excel file.
#
# Parameters in:
#	@ws: 		The excel file that you want to remove dupplicated rows.
#	@out_file: 	The output file that was removed dupplicated rows.
#
# Return:	NULL
# =============================================================================
def remove_dupplicate_rows(ws, out_file):

	data = pd.read_excel(ws, engine="openpyxl")
	df = data.drop_duplicates()
	df.to_excel(out_file)  



# =============================================================================
# This function is used to get all html source of website.
#
# Parameters in:
#	@site: 		Address of website that you want to get.
#	@dir :		Directory that you wamt to save the source files.
#
# Return:	NULL
# =============================================================================
def get_source_file(site, dir):

	os.system("cd " + dir + ";wget -r -nd --reject jpg,png "+ site +"; cd -")


# =============================================================================
# This function perform all process to get data from websites.
# https://concung.com/
# https://bibomart.com.vn/
# https://www.kidsplaza.vn/
# https://shoptretho.com.vn/
# https://www.tuticare.com/
# https://www.bachhoaxanh.com/
# https://www.kiotviet.vn/
#
# Parameters in:
#	@name_of_shop: 	Name of shop (KiotViet, ConCung, ConCung, ...).
#	@action :		Action to do (all, get_data, remove_dup).
#
# Return:	NULL
# =============================================================================
def main_scraping(name_of_shop, action):
	print("============= "+ name_of_shop +" ===============")

	if action == "get_source" or action == "all":
		# Get data from files, then write to excel file.
		if name_of_shop == "KiotViet":
			# Get source files from https://www.kiotviet.vn/
			get_source_file("https://www.kiotviet.vn/",source_dir_KiotViet_1)
		elif name_of_shop == "ConCung":
			# Get source files from https://concung.com/
			get_source_file("https://concung.com/",source_dir_ConCung_1)
		elif name_of_shop == "BiBo":
			# Get source files from https://bibomart.com.vn/
			get_source_file("https://bibomart.com.vn/",source_dir_BiBo_1)
		elif name_of_shop == "KidsPlaza":
			# Get source files from https://www.kidsplaza.vn/
			get_source_file("https://www.kidsplaza.vn/",source_dir_KidsPlaza_1)
		elif name_of_shop == "ShopTreTho":
			# Get source files from https://shoptretho.com.vn/
			get_source_file("https://shoptretho.com.vn/",source_dir_ShopTreTho_1)
		elif name_of_shop == "TutiCare":
			# Get source files from https://www.tuticare.com/
			get_source_file("https://www.tuticare.com/",source_dir_TutiCare_1)
		elif name_of_shop == "BXH":
			# Get source files from https://www.bachhoaxanh.com/
			get_source_file("https://www.bachhoaxanh.com/",source_dir_BXH_1)
		else:
			print("ERROR : Wrong name of shop")
			sys.exit()

	if action == "process_data" or action == "get_data" or action == "all":
		# Get list of files from source directory.
		list_file = get_list_of_file(Work_Space_0 +"\\example_"+ name_of_shop)
		print(list_file)

		# Get data from files, then write to excel file.
		if name_of_shop == "KiotViet":
			KiotViet_get_and_write_data(list_file, name_of_shop + ".xlsx")
		elif name_of_shop == "ConCung":
			ConCung_get_and_write_data(list_file, name_of_shop + ".xlsx")
		elif name_of_shop == "BiBo":
			BiBo_get_and_write_data(list_file, name_of_shop + ".xlsx")
		elif name_of_shop == "KidsPlaza":
			KidsPlaza_get_and_write_data(list_file, name_of_shop + ".xlsx")
		elif name_of_shop == "ShopTreTho":
			ShopTreTho_get_and_write_data(list_file, name_of_shop + ".xlsx")
		elif name_of_shop == "TutiCare":
			TutiCare_get_and_write_data(list_file, name_of_shop + ".xlsx")
		elif name_of_shop == "BXH":
			BXH_get_and_write_data(list_file, name_of_shop + ".xlsx")
		else:
			print("ERROR : Wrong name of shop")
			sys.exit()
	if action == "process_data" or action == "remove_dup" or action == "all":
		# input file that is needed to remove dupplicated rows.
		in_file = Work_Space_0 + "\\" + name_of_shop + ".xlsx"

		# output file that is removed dupplicated rows.
		out_file = Work_Space_0 + "\\" + name_of_shop + "_Innovation.xlsx"

		remove_dupplicate_rows(in_file, out_file)


def main(argv):

	# ================ Print help and get options ===================
	try:
		opts, args = getopt.getopt(argv,"ha:s",["action=","shop="])
	except getopt.GetoptError:
		print("	./web_scraping.py --shop <1/2/3/4/5/6/7> --action <all/get_source/get_data/remove_dup>\n")
		sys.exit(2)
	for opt, arg in opts:
		if opt in ('-h', '--help'):
			print("Command:")
			print("	./web_scraping.py --shop <1,2,3,4,5,6,7> --action <all/get_source/remove_dup/process_data>\n")
			print("Mandatory arguments:\n")
			print("--shop:")
			print("	1. KiotViet | 2. ConCung | 3. BiBo | 4. KidsPlaza | 5. ShopTreTho | 6. TutiCare | 7. BXH")
			print("--action:")
			print("	all : Do all steps. \n\tget_source: Fetch source only.\n\tget_data: Get data from source write to excel only.")
			print("	remove_dup: Remove dupplicated data only.\n\tprocess_data: include get_data --> remove_dup")
			sys.exit()
		elif opt == '--shop':
			name_of_shop = arg
		elif opt == '--action':
			action = arg
		else:
			print("Wrong command line \nPlease follow below command:\n")
			print("	./web_scraping.py --shop <1/2/3/4/5/6/7> --action <all/get_source/get_data/remove_dup>\n")
			sys.exit(2)

	# =========== Get list of shop from arguments ===================
	list_of_shop_tmp = name_of_shop.split(',')
	list_of_shop = []

	for item in list_of_shop_tmp:
		if item == "1":
			list_of_shop.append("KiotViet")
		elif item == "2":
			list_of_shop.append("ConCung")
		elif item == "3":
			list_of_shop.append("BiBo")
		elif item == "4":
			list_of_shop.append("KidsPlaza")
		elif item == "5":
			list_of_shop.append("ShopTreTho")
		elif item == "6":
			list_of_shop.append("TutiCare")
		elif item == "7":
			list_of_shop.append("BXH")

	# ============================ Start the program ================
	print("================ Start Program =================")
	start = time.time()

	for item in list_of_shop:
		main_scraping(item, action)

	end = time.time()
	elapsed = int(end - start)
	print("================ End of Program ================")

	# Print the time duration of the program.
	print('DURATION : {:02d}:{:02d}:{:02d}'.format(elapsed // 3600, (elapsed % 3600 // 60), elapsed % 60))

if __name__ == "__main__":
	main(sys.argv[1:])